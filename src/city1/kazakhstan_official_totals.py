from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OFFICIAL_SOURCE_NAME = "Bureau of National Statistics of Kazakhstan - regional population spreadsheet"
OFFICIAL_REGION_PAGE_URL = "https://stat.gov.kz/ru/region/?param=POPULATION"


@dataclass(frozen=True)
class WorkbookCitySource:
    city_name: str
    normalized_city_name: str
    cache_filename: str
    source_url: str
    reference_date: str
    row_terms: tuple[str, ...]
    worksheet_name: str = "1"
    population_column_index: int = 5
    country: str = "Kazakhstan"
    source_tier: str = "official"
    verified: bool = True


MANUAL_VERIFIED_CITY_RECORDS: tuple[dict[str, Any], ...] = (
    {
        "city_name": "Almaty",
        "normalized_city_name": "almaty",
        "country": "Kazakhstan",
        "population": 2351424,
        "reference_date": "2026-02-01",
        "source_tier": "official",
        "verified": True,
        "source_name": "Bureau of National Statistics of Kazakhstan - Region statistics",
        "source_url": OFFICIAL_REGION_PAGE_URL,
        "notes": "Official value from the BNS regional statistics page, as of 2026-02-01.",
    },
    {
        "city_name": "Astana",
        "normalized_city_name": "astana",
        "country": "Kazakhstan",
        "population": 1649242,
        "reference_date": "2026-02-01",
        "source_tier": "official",
        "verified": True,
        "source_name": "Bureau of National Statistics of Kazakhstan - Region statistics",
        "source_url": OFFICIAL_REGION_PAGE_URL,
        "notes": "Official value from the BNS regional statistics page, as of 2026-02-01.",
    },
    {
        "city_name": "Shymkent",
        "normalized_city_name": "shymkent",
        "country": "Kazakhstan",
        "population": 1298279,
        "reference_date": "2026-02-01",
        "source_tier": "official",
        "verified": True,
        "source_name": "Bureau of National Statistics of Kazakhstan - Region statistics",
        "source_url": OFFICIAL_REGION_PAGE_URL,
        "notes": "Official value from the BNS regional statistics page, as of 2026-02-01.",
    },
)


WORKBOOK_CITY_SOURCES: tuple[WorkbookCitySource, ...] = (
    WorkbookCitySource(
        city_name="Semey",
        normalized_city_name="semey",
        cache_filename="abay_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477499/file/en/",
        reference_date="2026-02-01",
        row_terms=("Semey city administration",),
    ),
    WorkbookCitySource(
        city_name="Kurchatov",
        normalized_city_name="kurchatov",
        cache_filename="abay_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477499/file/en/",
        reference_date="2026-02-01",
        row_terms=("Kurchatov city administration",),
    ),
    WorkbookCitySource(
        city_name="Taraz",
        normalized_city_name="taraz",
        cache_filename="zhambyl_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477540/file/en/",
        reference_date="2026-02-01",
        row_terms=("Taraz city",),
    ),
    WorkbookCitySource(
        city_name="Uralsk",
        normalized_city_name="uralsk",
        cache_filename="zko_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477560/file/en/",
        reference_date="2026-02-01",
        row_terms=("Uralsk c.",),
    ),
    WorkbookCitySource(
        city_name="Petropavlovsk",
        normalized_city_name="petropavlovsk",
        cache_filename="sko_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477495/file/ru/",
        reference_date="2026-02-01",
        row_terms=("Петропавловск",),
    ),
    WorkbookCitySource(
        city_name="Ust Kamenogorsk",
        normalized_city_name="ust kamenogorsk",
        cache_filename="vko_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477478/file/en/",
        reference_date="2026-02-01",
        row_terms=("Ust-Kamenogorsk city administration",),
    ),
    WorkbookCitySource(
        city_name="Ridder",
        normalized_city_name="ridder",
        cache_filename="vko_2026_02_01.xlsx",
        source_url="https://stat.gov.kz/api/iblock/element/region/477478/file/en/",
        reference_date="2026-02-01",
        row_terms=("Ridder city administration",),
    ),
)


def _coerce_positive_integer(value: object) -> int | None:
    if value is None:
        return None

    try:
        population = int(round(float(value)))
    except (TypeError, ValueError):
        return None

    if population <= 0:
        return None
    return population


def _extract_city_population(source: WorkbookCitySource, cache_dir: Path) -> tuple[int, str]:
    workbook_path = cache_dir / source.cache_filename
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook is missing: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[source.worksheet_name] if source.worksheet_name in workbook.sheetnames else workbook[workbook.sheetnames[-1]]

    for row in worksheet.iter_rows(values_only=True):
        label = row[0]
        if not isinstance(label, str):
            continue

        lower_label = label.lower()
        if not any(term.lower() in lower_label for term in source.row_terms):
            continue

        population = _coerce_positive_integer(row[source.population_column_index])
        if population is None:
            raise ValueError(
                f"Matched row for {source.city_name!r} but population column "
                f"{source.population_column_index} was invalid in {workbook_path.name}."
            )

        return population, label.strip()

    raise ValueError(f"Could not find a matching row for {source.city_name!r} in {workbook_path.name}.")


def build_verified_kazakhstan_city_records(cache_dir: str | Path) -> tuple[list[dict[str, Any]], tuple[str, ...]]:
    records: list[dict[str, Any]] = [dict(record) for record in MANUAL_VERIFIED_CITY_RECORDS]
    warnings: list[str] = []
    tables_dir = Path(cache_dir)

    for source in WORKBOOK_CITY_SOURCES:
        try:
            population, matched_label = _extract_city_population(source, tables_dir)
        except Exception as exc:
            warnings.append(f"{source.city_name}: {exc}")
            continue

        records.append(
            {
                "city_name": source.city_name,
                "normalized_city_name": source.normalized_city_name,
                "country": source.country,
                "population": population,
                "reference_date": source.reference_date,
                "source_tier": source.source_tier,
                "verified": source.verified,
                "source_name": OFFICIAL_SOURCE_NAME,
                "source_url": source.source_url,
                "notes": (
                    "Official city-level value extracted from a BNS regional monthly population "
                    f"spreadsheet, matched row {matched_label!r}, as of {source.reference_date}."
                ),
            }
        )

    return records, tuple(warnings)
